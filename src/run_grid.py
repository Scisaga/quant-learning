from __future__ import annotations

"""
批量实验网格（Grid Runner）：并行跑“训练 + 回测报告”，并把每个实验的产物路径记录下来。

你提出的对比维度，本脚本覆盖：
1) 市场维度：all、csi300、csi1000（可通过 --markets 自定义）
2) label 维度：horizon ∈ {1,5,10}（可通过 --label-horizons 自定义）
3) PIT 维度：默认对比“不含 PIT vs 含 PIT（4 个字段一起）”
   - 默认：--pit-grid all  -> no_pit + pit_all
   - 可选：--pit-grid single / all+single  -> 增加单字段 PIT 变体
4) 时间维度：从 2010 年滚动到最新时间，walk-forward 窗口（训练/验证/测试按年滚动）

并行化策略：
- 这里用 ThreadPoolExecutor 并行跑“外部子进程”：
  - `python src/train/train_lgb_alpha158_pit.py ...`
  - `python src/backtest/generate_html_report.py ...`
- 每个 job 之间互不共享 Qlib 的全局状态（子进程隔离），因此线程并行是安全的、实现也最简单。

输出（默认在 reports/grid_runs/<timestamp>/）：
- summary.json：本次展开后的窗口与网格参数
- results.jsonl：每个 job 一行（包含 recorder_id、HTML 报告路径、日志路径、状态）
- logs/*.log：训练/报告的 stdout+stderr 合并日志

使用建议：
- 先 `--dry-run` 看任务总量，默认会非常多（市场*label*PIT*窗口）。
- 建议先缩小 markets/label 或减少窗口长度，在确认数据/流程稳定后再放开并行与任务量。
"""

import argparse
import csv
import json
import os
import re
import subprocess
import sys
import threading
from dataclasses import asdict, dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Optional

# tqdm 进度条：缺失时自动降级为纯 print
try:
    from tqdm.auto import tqdm  # type: ignore
except Exception:  # pragma: no cover
    tqdm = None  # type: ignore


# 解析 report.log 中关键指标（risk_analysis / IC）
_RISK_METRIC_LINE_RE = re.compile(
    r"^(?P<k>mean|std|annualized_return|information_ratio|max_drawdown)\s+(?P<v>[-+0-9.eEinfNaN]+)\s*$",
    flags=re.IGNORECASE,
)
_IC_BLOCK_RE = re.compile(r"\{\s*'IC'\s*:\s*[^}]+\}", flags=re.DOTALL)
_IC_KV_RE = re.compile(
    r"'(?P<k>IC|ICIR|Rank IC|Rank ICIR)'\s*:\s*(?:np\.float64\()?\s*(?P<v>[-+0-9.eEinfNaN]+)\s*\)?"
)


def _comma_list(s: str) -> list[str]:
    """把逗号分隔的字符串解析成 list，并去掉空白项。"""
    return [x.strip() for x in s.split(",") if x.strip()]


def _parse_date(s: str) -> date:
    """解析 YYYY-MM-DD 到 date。"""
    return datetime.strptime(s, "%Y-%m-%d").date()


def _date_str(d: date) -> str:
    """date -> YYYY-MM-DD 字符串。"""
    return d.isoformat()


def _year_start(y: int) -> date:
    """某年的 1 月 1 日。"""
    return date(y, 1, 1)


def _year_end(y: int) -> date:
    """某年的 12 月 31 日。"""
    return date(y, 12, 31)


def _slug(s: str) -> str:
    """把任意字符串转成更适合做文件名/标识的 slug。"""
    s = s.strip().lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    return s.strip("_") or "x"


def _benchmark_for_market(market: str) -> str:
    """
    给定 market 返回默认 benchmark。

    注意：
    - “all” 并不是一个严格指数成分集合，默认用 SH000300 作为基准；
    - 若你有更合适的 benchmark（比如全市场用某个宽基），请用 --benchmark-map 覆盖。
    """
    market = market.lower()
    if market == "csi1000":
        return "SH000852"
    if market == "csi300":
        return "SH000300"
    return "SH000300"


def _iter_walk_forward_windows(
    *,
    start: date,
    end: date,
    train_years: int,
    valid_years: int,
    test_years: int,
    step_years: int,
    allow_partial_last_window: bool,
    min_test_days: int,
) -> list["Window"]:
    """
    生成 walk-forward 窗口列表（按年滚动）。

    默认窗口语义（可通过参数改）：
    - 训练 train_years 年
    - 验证 valid_years 年
    - 测试 test_years 年
    - 每次向前滚 step_years 年

    handler_range 的 start/end：
    - 为了避免数据切片缺失，这里把 handler 覆盖到 train_start..test_end（同一个窗口的全周期）。
    """
    if train_years <= 0 or valid_years <= 0 or test_years <= 0 or step_years <= 0:
        raise ValueError("train_years/valid_years/test_years/step_years must be positive")

    windows: list[Window] = []
    cursor_year = start.year

    while True:
        train_start = max(start, _year_start(cursor_year))
        train_end = _year_end(cursor_year + train_years - 1)
        valid_start = _year_start(train_end.year + 1)
        valid_end = _year_end(valid_start.year + valid_years - 1)
        test_start = _year_start(valid_end.year + 1)
        test_end_full = _year_end(test_start.year + test_years - 1)
        if test_start > end:
            break

        test_end = min(end, test_end_full)
        if not allow_partial_last_window and test_end < test_end_full:
            break

        if test_end < test_start:
            break

        if (test_end - test_start).days + 1 < min_test_days:
            break

        handler_start = train_start
        handler_end = test_end
        windows.append(
            Window(
                handler_start=_date_str(handler_start),
                handler_end=_date_str(handler_end),
                train_start=_date_str(train_start),
                train_end=_date_str(train_end),
                valid_start=_date_str(valid_start),
                valid_end=_date_str(valid_end),
                test_start=_date_str(test_start),
                test_end=_date_str(test_end),
            )
        )

        cursor_year += step_years

    return windows


@dataclass(frozen=True)
class Window:
    """单个 walk-forward 窗口的时间配置。"""
    handler_start: str
    handler_end: str
    train_start: str
    train_end: str
    valid_start: str
    valid_end: str
    test_start: str
    test_end: str

    @property
    def tag(self) -> str:
        """用于 job 唯一标识与日志文件名的一部分。"""
        return f"{self.train_start}_to_{self.test_end}"


@dataclass(frozen=True)
class PitSpec:
    """PIT 维度的一种设置：名字 + 传给训练脚本的 CLI 参数片段。"""
    name: str
    pit_fields_arg: list[str]


@dataclass(frozen=True)
class JobSpec:
    """一个最小实验单元：市场 x label x PIT x 时间窗口。"""
    market: str
    benchmark: str
    label_horizon: int
    label_expr: str
    pit: PitSpec
    window: Window
    exp_name: str

    @property
    def tag(self) -> str:
        """job 的稳定 key（用于 results.jsonl 里 resume 跳过）。"""
        return _slug(f"{self.market}_h{self.label_horizon}_{self.pit.name}_{self.window.tag}")


@dataclass
class JobResult:
    """单个 job 的执行结果（用于落到 results.jsonl）。"""
    job: JobSpec
    status: str
    started_at: str
    finished_at: str
    train_returncode: Optional[int] = None
    report_returncode: Optional[int] = None
    recorder_id: Optional[str] = None
    report_html: Optional[str] = None
    error: Optional[str] = None
    train_log: Optional[str] = None
    report_log: Optional[str] = None


RECORDER_RE = re.compile(r"recorder_id=([0-9a-fA-F]+)")
# 注意：不要用 `$` 锚定“整个输出的末尾”，因为 Qlib/plotly 可能在 `written:` 之后还有额外输出。
WRITTEN_RE = re.compile(r"written:\s*(.+\.html)\s*$", flags=re.MULTILINE)


def _utc_now_iso() -> str:
    """返回 UTC 的 ISO 时间戳（避免 Python 3.12+ 对 utcnow 的弃用警告）。"""
    return datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")


def _run_cmd(*, cmd: list[str], cwd: Path, log_path: Path, env: dict[str, str]) -> tuple[int, str]:
    """运行外部命令，把 stdout+stderr 合并写到 log_path，并返回 (returncode, output_text)。"""
    p = subprocess.run(
        cmd,
        cwd=str(cwd),
        env=env,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
    )
    out = p.stdout or ""
    log_path.parent.mkdir(parents=True, exist_ok=True)
    log_path.write_text(out, encoding="utf-8")
    return p.returncode, out


def _extract_recorder_id(output: str) -> Optional[str]:
    """从训练脚本输出里提取 `recorder_id=...`（取最后一个匹配）。"""
    m = None
    for m in RECORDER_RE.finditer(output):
        pass
    return m.group(1) if m else None


def _extract_written_html(output: str) -> Optional[str]:
    """从报告脚本输出里提取 `written: <path>.html`（取最后一个匹配）。"""
    last: Optional[str] = None
    for m in WRITTEN_RE.finditer(output):
        last = m.group(1).strip()
    return last


def _run_one_job(
    *,
    job: JobSpec,
    repo_root: Path,
    python_exe: str,
    provider_uri: str,
    lgb_num_threads: Optional[int],
    topk: int,
    n_drop: int,
    hold_thresh: int,
    account: float,
    open_cost: float,
    close_cost: float,
    min_cost: float,
    limit_threshold: float,
    out_dir: Path,
    pit_missing: str,
    joblib_backend: str,
) -> JobResult:
    """
    执行单个 job：训练 -> 生成 HTML 报告。

    关键点：
    - 训练脚本会创建 Qlib recorder，并在 stdout 打印 recorder_id；
    - 报告脚本用 (exp_name, recorder_id) 定位 recorder，生成/补齐回测与分析产物，再写 HTML；
    - 这里强制 backtest 区间对齐 test 段，保证口径一致。
    """
    started_at = _utc_now_iso()
    logs_dir = out_dir / "logs"
    train_log = logs_dir / f"{job.tag}.train.log"
    report_log = logs_dir / f"{job.tag}.report.log"

    env = dict(os.environ)
    env["PYTHONUNBUFFERED"] = "1"

    train_cmd = [
        python_exe,
        "src/train/train_lgb_alpha158_pit.py",
        "--provider-uri",
        provider_uri,
        "--market",
        job.market,
        "--benchmark",
        job.benchmark,
        "--exp-name",
        job.exp_name,
        "--start-time",
        job.window.handler_start,
        "--end-time",
        job.window.handler_end,
        "--train",
        f"{job.window.train_start},{job.window.train_end}",
        "--valid",
        f"{job.window.valid_start},{job.window.valid_end}",
        "--test",
        f"{job.window.test_start},{job.window.test_end}",
        "--label-expr",
        job.label_expr,
        "--joblib-backend",
        joblib_backend,
    ]
    train_cmd.extend(job.pit.pit_fields_arg)
    train_cmd.extend(["--pit-missing", pit_missing])
    if lgb_num_threads is not None:
        train_cmd.extend(["--lgb-num-threads", str(lgb_num_threads)])

    rc_train, out_train = _run_cmd(cmd=train_cmd, cwd=repo_root, log_path=train_log, env=env)
    recorder_id = _extract_recorder_id(out_train)
    if rc_train != 0 or not recorder_id:
        finished_at = _utc_now_iso()
        return JobResult(
            job=job,
            status="train_failed",
            started_at=started_at,
            finished_at=finished_at,
            train_returncode=rc_train,
            recorder_id=recorder_id,
            error="training failed or recorder_id not found in output",
            train_log=str(train_log),
        )

    report_cmd = [
        python_exe,
        "src/backtest/generate_html_report.py",
        "--provider-uri",
        provider_uri,
        "--exp-name",
        job.exp_name,
        "--recorder-id",
        recorder_id,
        "--benchmark",
        job.benchmark,
        "--backtest-start",
        job.window.test_start,
        "--backtest-end",
        job.window.test_end,
        "--topk",
        str(topk),
        "--n-drop",
        str(n_drop),
        "--hold-thresh",
        str(hold_thresh),
        "--account",
        str(account),
        "--open-cost",
        str(open_cost),
        "--close-cost",
        str(close_cost),
        "--min-cost",
        str(min_cost),
        "--limit-threshold",
        str(limit_threshold),
    ]
    rc_report, out_report = _run_cmd(cmd=report_cmd, cwd=repo_root, log_path=report_log, env=env)
    html_path = _extract_written_html(out_report)

    finished_at = _utc_now_iso()
    status = "ok" if rc_report == 0 and html_path else "report_failed"
    err: Optional[str] = None
    if status != "ok":
        if rc_report != 0:
            err = f"report command failed (returncode={rc_report}); see report_log"
        else:
            err = "report command returned 0 but HTML path was not detected in output; see report_log"
    return JobResult(
        job=job,
        status=status,
        started_at=started_at,
        finished_at=finished_at,
        train_returncode=rc_train,
        report_returncode=rc_report,
        recorder_id=recorder_id,
        report_html=html_path,
        error=err,
        train_log=str(train_log),
        report_log=str(report_log),
    )


def _load_completed_keys(path: Path) -> set[str]:
    """
    从 results.jsonl 中加载 status=ok 的 key，用于 --resume 跳过已完成任务。
    """
    if not path.exists():
        return set()
    done: set[str] = set()
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            obj = json.loads(line)
            if obj.get("status") == "ok":
                key = obj.get("key")
                if isinstance(key, str):
                    done.add(key)
        except Exception:
            continue
    return done


def _to_number_or_text(s: str) -> Any:
    """尽量把字符串转成数字（float/int），失败则返回原文本。"""
    t = s.strip()
    if t == "":
        return ""
    low = t.lower()
    if low in {"nan", "none", "null"}:
        return ""
    try:
        # 兼容 1,234 这种展示
        tt = t.replace(",", "")
        f = float(tt)
        if f.is_integer():
            return int(f)
        return f
    except Exception:
        return t


def _extract_metrics_from_report_log(report_log_path: Path) -> dict[str, Any]:
    """
    从 generate_html_report.py 的 stdout 日志中提取关键指标：
    - excess return without/with cost：mean/std/annualized_return/information_ratio/max_drawdown
    - IC/ICIR/RankIC/RankICIR（SigAnaRecord 输出的字典）
    """
    if not report_log_path.exists():
        return {}
    txt = report_log_path.read_text(encoding="utf-8", errors="replace")
    out: dict[str, Any] = {}

    # 1) risk blocks (scan line-by-line; robust to quotes / extra header rows like `risk`)
    prefix: Optional[str] = None
    for line in txt.splitlines():
        low = line.lower()
        if "analysis results of the excess return without cost(1day)" in low:
            prefix = "excess_return_without_cost"
            continue
        if "analysis results of the excess return with cost(1day)" in low:
            prefix = "excess_return_with_cost"
            continue
        if prefix is None:
            continue

        # End current block when another analysis section begins
        if "analysis results of" in low and "excess return" not in low:
            prefix = None
            continue

        stripped = line.strip()
        if stripped.lower() == "risk":
            continue

        m = _RISK_METRIC_LINE_RE.match(stripped)
        if m:
            k = (m.group("k") or "").strip().lower()
            v = _to_number_or_text(m.group("v") or "")
            out[f"{prefix}.{k}"] = v

    # 2) IC dict (take the last one)
    ic_block = None
    for m in _IC_BLOCK_RE.finditer(txt):
        ic_block = m.group(0)
    if ic_block:
        kvs: dict[str, Any] = {}
        for km in _IC_KV_RE.finditer(ic_block):
            k = (km.group("k") or "").strip()
            v = _to_number_or_text(km.group("v") or "")
            kvs[k] = v
        if "IC" in kvs:
            out["IC"] = kvs.get("IC")
        if "ICIR" in kvs:
            out["ICIR"] = kvs.get("ICIR")
        if "Rank IC" in kvs:
            out["RankIC"] = kvs.get("Rank IC")
        if "Rank ICIR" in kvs:
            out["RankICIR"] = kvs.get("Rank ICIR")

    return out


def _parse_kv_table(table) -> dict[str, str]:
    """
    解析 generate_html_report.py 里 `table.kv` 的键值对。

    结构示例：
    <table class='kv'><tbody>
      <tr class='kv-section'><td colspan='2'>Run</td></tr>
      <tr><td>experiment_name</td><td>xxx</td></tr>
    </tbody></table>
    """
    out: dict[str, str] = {}
    for tr in table.find_all("tr"):
        if "kv-section" in (tr.get("class") or []):
            continue
        tds = tr.find_all("td")
        if len(tds) != 2:
            continue
        k = tds[0].get_text(" ", strip=True)
        v = tds[1].get_text(" ", strip=True)
        if k:
            out[k] = v
    return out


def _parse_html_table(table) -> tuple[list[str], list[list[str]]]:
    """把 HTML table 解析成 (headers, rows)；只处理本项目生成的常见结构。"""
    headers: list[str] = []
    thead = table.find("thead")
    if thead is not None:
        hr = thead.find("tr")
        if hr is not None:
            headers = [th.get_text(" ", strip=True) for th in hr.find_all("th")]
    tbody = table.find("tbody")
    trs = (tbody or table).find_all("tr")
    rows: list[list[str]] = []
    for tr in trs:
        cells = tr.find_all(["th", "td"])
        if not cells:
            continue
        rows.append([c.get_text(" ", strip=True) for c in cells])
    # 如果有 thead，则 tbody 的 tr 才是真正数据；上面的 rows 可能把 thead 也读进来（不同生成器不一致）
    # 这里做一个简单去重：如果第一行等于 headers，则丢掉第一行。
    if headers and rows and rows[0] == headers:
        rows = rows[1:]
    return headers, rows


def _extract_params_from_report_html(html_path: Path) -> dict[str, str]:
    """从单份报告 HTML 提取 Experiment Parameters（kv 表），返回原始 key->value（字符串）。"""
    try:
        from bs4 import BeautifulSoup  # type: ignore
    except ModuleNotFoundError as e:
        raise RuntimeError("missing dependency: beautifulsoup4 (bs4)") from e

    text = html_path.read_text(encoding="utf-8", errors="replace")
    soup = BeautifulSoup(text, "html.parser")

    out: dict[str, str] = {}
    h2_params = soup.find("h2", string="Experiment Parameters")
    if h2_params is not None:
        table = h2_params.find_next("table", attrs={"class": "kv"})
        if table is not None:
            out.update(_parse_kv_table(table))
    return out


def _pick_params(params: dict[str, str]) -> dict[str, Any]:
    """
    从 Experiment Parameters（kv）里挑选“有区分度、且对评估有帮助”的字段。
    避免重复（如 label_expr）、避免无信息字段（provider_uri/train_script 等）。
    """
    keep = [
        "pit_fields",
        # Prediction coverage / sample size
        "pred_coverage",
        "pred_instruments",
        # Model
        "model_class",
        "num_leaves",
        "max_depth",
        "learning_rate",
        "subsample",
        "colsample_bytree",
        "lambda_l1",
        "lambda_l2",
        "num_threads",
        # Strategy
        "strategy.topk",
        "strategy.n_drop",
        "strategy.hold_thresh",
        # Exchange/cost
        "exchange.open_cost",
        "exchange.close_cost",
        "exchange.min_cost",
        "exchange.limit_threshold",
    ]
    out: dict[str, Any] = {}
    for k in keep:
        if k in params:
            out[k] = _to_number_or_text(params[k])
    # pred_instruments: "348 unique" -> 348
    pi = params.get("pred_instruments", "")
    if isinstance(pi, str):
        m = re.match(r"\s*(\d+)\b", pi)
        if m:
            out["pred_instruments"] = int(m.group(1))
    # Split pred_coverage
    cov = params.get("pred_coverage", "")
    if isinstance(cov, str) and " .. " in cov:
        a, b = cov.split(" .. ", 1)
        out["pred_start"] = a.strip()
        out["pred_end"] = b.strip()
    return out


def _export_compare_table(
    *,
    repo_root: Path,
    out_dir: Path,
    results_path: Path,
    progress: bool,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """读取 results.jsonl，返回 (ok_rows, error_rows)。"""
    ok_rows: list[dict[str, Any]] = []
    err_rows: list[dict[str, Any]] = []
    if not results_path.exists():
        return ok_rows, err_rows

    total = None
    if progress and tqdm is not None:
        try:
            with results_path.open("r", encoding="utf-8") as f:
                total = sum(1 for _ in f)
        except Exception:
            total = None

    it = results_path.open("r", encoding="utf-8")
    try:
        line_iter = it
        if progress and tqdm is not None:
            line_iter = tqdm(it, total=total, desc="export rows", dynamic_ncols=True)  # type: ignore[assignment]

        for line in line_iter:
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
            except Exception:
                continue

            status = obj.get("status", "")
            window = obj.get("window") if isinstance(obj.get("window"), dict) else {}

            base: dict[str, Any] = {
                "key": obj.get("key", ""),
                "exp_name": obj.get("exp_name", ""),
                "market": obj.get("market", ""),
                "benchmark": obj.get("benchmark", ""),
                "label_horizon": obj.get("label_horizon", ""),
                "label_expr": obj.get("label_expr", ""),
                "pit": obj.get("pit", ""),
                "window.train_start": window.get("train_start", "") if isinstance(window, dict) else "",
                "window.train_end": window.get("train_end", "") if isinstance(window, dict) else "",
                "window.valid_start": window.get("valid_start", "") if isinstance(window, dict) else "",
                "window.valid_end": window.get("valid_end", "") if isinstance(window, dict) else "",
                "window.test_start": window.get("test_start", "") if isinstance(window, dict) else "",
                "window.test_end": window.get("test_end", "") if isinstance(window, dict) else "",
                "recorder_id": obj.get("recorder_id", ""),
                "report_html": obj.get("report_html", ""),
            }

            if status != "ok":
                err_rows.append(
                    {
                        "key": base["key"],
                        "status": status,
                        "error": obj.get("error", ""),
                        "train_log": obj.get("train_log", ""),
                        "report_log": obj.get("report_log", ""),
                    }
                )
                continue

            row = dict(base)

            # 1) metrics from report log (most robust for grid summary)
            report_log = obj.get("report_log")
            if isinstance(report_log, str) and report_log.strip():
                p = Path(report_log.strip())
                report_log_path = p if p.is_absolute() else (repo_root / p)
                row.update(_extract_metrics_from_report_log(report_log_path))

            # 2) selected params from HTML
            report_html = obj.get("report_html")
            if isinstance(report_html, str) and report_html.strip():
                p = Path(report_html.strip())
                html_path = p if p.is_absolute() else (repo_root / p)
                if html_path.exists():
                    try:
                        params = _extract_params_from_report_html(html_path)
                        row.update(_pick_params(params))
                    except Exception:
                        pass

            ok_rows.append(row)
    finally:
        it.close()

    return ok_rows, err_rows


def _write_csv(path: Path, *, rows: list[dict[str, Any]], columns: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, "") for k in columns})


def _export_compare_xlsx(
    *,
    out_path: Path,
    rows: list[dict[str, Any]],
    columns: list[str],
    err_rows: list[dict[str, Any]],
    progress: bool,
) -> None:
    """
    导出 XLSX（比 CSV 更适合做横向对比）：
    - Summary：核心对比表（冻结首行 + 自动筛选 + 合理列宽 + 数字格式）
    - Errors：失败任务列表（用于排查）
    """
    try:
        from openpyxl import Workbook  # type: ignore
        from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore
        from openpyxl.utils import get_column_letter  # type: ignore
    except ModuleNotFoundError as e:
        raise RuntimeError("missing dependency: openpyxl") from e

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    header_fill = PatternFill("solid", fgColor="F2F2F2")
    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap_align = Alignment(vertical="top", wrap_text=True)

    # Write header
    ws.append(columns)
    for c, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    # Write rows (with optional tqdm)
    row_iter = rows
    if progress and tqdm is not None:
        row_iter = tqdm(rows, desc="write xlsx", dynamic_ncols=True)  # type: ignore[assignment]

    for r in row_iter:
        ws.append([r.get(col, "") for col in columns])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"

    # Number formats
    percent_cols = {
        "excess_return_without_cost.mean",
        "excess_return_without_cost.std",
        "excess_return_without_cost.annualized_return",
        "excess_return_without_cost.max_drawdown",
        "excess_return_with_cost.mean",
        "excess_return_with_cost.std",
        "excess_return_with_cost.annualized_return",
        "excess_return_with_cost.max_drawdown",
    }
    int_cols = {
        "label_horizon",
        "pred_instruments",
        "num_leaves",
        "max_depth",
        "num_threads",
        "strategy.topk",
        "strategy.n_drop",
        "strategy.hold_thresh",
    }
    float_cols = {
        "excess_return_without_cost.information_ratio",
        "excess_return_with_cost.information_ratio",
        "IC",
        "ICIR",
        "RankIC",
        "RankICIR",
        "learning_rate",
        "subsample",
        "colsample_bytree",
        "lambda_l1",
        "lambda_l2",
        "exchange.open_cost",
        "exchange.close_cost",
        "exchange.min_cost",
        "exchange.limit_threshold",
    }

    for j, col in enumerate(columns, start=1):
        fmt = None
        if col in percent_cols:
            fmt = "0.00%"
        elif col in int_cols:
            fmt = "0"
        elif col in float_cols or col.startswith("excess_return_"):
            fmt = "0.0000"
        if fmt:
            for i in range(2, ws.max_row + 1):
                ws.cell(row=i, column=j).number_format = fmt
        # Wrap some long-text columns
        if col in {"label_expr", "report_html"}:
            for i in range(2, ws.max_row + 1):
                ws.cell(row=i, column=j).alignment = wrap_align

    # Column widths (cap to avoid extreme wide columns)
    for j, col in enumerate(columns, start=1):
        max_len = len(col)
        for i in range(2, min(ws.max_row, 2000) + 1):
            v = ws.cell(row=i, column=j).value
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        width = min(max(10, max_len + 2), 60)
        ws.column_dimensions[get_column_letter(j)].width = width

    # Errors sheet
    if err_rows:
        ws2 = wb.create_sheet("Errors")
        err_cols = ["key", "status", "error", "train_log", "report_log"]
        ws2.append(err_cols)
        for c in range(1, len(err_cols) + 1):
            cell = ws2.cell(row=1, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
        err_iter = err_rows
        if progress and tqdm is not None:
            err_iter = tqdm(err_rows, desc="write errors", dynamic_ncols=True)  # type: ignore[assignment]
        for r in err_iter:
            ws2.append([r.get(k, "") for k in err_cols])
        ws2.freeze_panes = "A2"
        ws2.auto_filter.ref = f"A1:{get_column_letter(len(err_cols))}1"
        for j, col in enumerate(err_cols, start=1):
            width = 16 if col in {"status"} else 60 if col in {"error", "train_log", "report_log"} else 32
            ws2.column_dimensions[get_column_letter(j)].width = width
            for i in range(2, ws2.max_row + 1):
                ws2.cell(row=i, column=j).alignment = wrap_align

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def _export_compare_csv(*, repo_root: Path, out_dir: Path, results_path: Path, progress: bool) -> Path:
    """
    从 results.jsonl + HTML 报告中提取对比字段，输出一个宽表 CSV（每个 job 一行）。
    """
    ok_rows, err_rows = _export_compare_table(repo_root=repo_root, out_dir=out_dir, results_path=results_path, progress=progress)

    # Summary columns (ordered, no duplicates/no noisy fields)
    cols = [
        "key",
        "exp_name",
        "market",
        "label_horizon",
        "label_expr",
        "pit",
        "pit_fields",
        "benchmark",
        "window.train_start",
        "window.train_end",
        "window.valid_start",
        "window.valid_end",
        "window.test_start",
        "window.test_end",
        "excess_return_without_cost.mean",
        "excess_return_without_cost.std",
        "excess_return_without_cost.annualized_return",
        "excess_return_without_cost.information_ratio",
        "excess_return_without_cost.max_drawdown",
        "excess_return_with_cost.mean",
        "excess_return_with_cost.std",
        "excess_return_with_cost.annualized_return",
        "excess_return_with_cost.information_ratio",
        "excess_return_with_cost.max_drawdown",
        "IC",
        "ICIR",
        "RankIC",
        "RankICIR",
        "pred_instruments",
        "pred_start",
        "pred_end",
        "model_class",
        "num_leaves",
        "max_depth",
        "learning_rate",
        "subsample",
        "colsample_bytree",
        "lambda_l1",
        "lambda_l2",
        "num_threads",
        "strategy.topk",
        "strategy.n_drop",
        "strategy.hold_thresh",
        "exchange.open_cost",
        "exchange.close_cost",
        "exchange.min_cost",
        "exchange.limit_threshold",
        "recorder_id",
        "report_html",
    ]

    out_csv = out_dir / "grid_compare.csv"
    _write_csv(out_csv, rows=ok_rows, columns=cols)

    # Errors table (separate file for排查)
    if err_rows:
        err_csv = out_dir / "grid_errors.csv"
        _write_csv(err_csv, rows=err_rows, columns=["key", "status", "error", "train_log", "report_log"])

    return out_csv


def main(argv: Optional[list[str]] = None) -> int:
    p = argparse.ArgumentParser(description="Run a market/label/PIT/time walk-forward grid (train + HTML report) in parallel.")
    p.add_argument("--provider-uri", default="data/qlib_data/cn_data")
    # 市场列表：直接透传给训练脚本里的 `D.instruments(market)`。
    # 常用：all / csi300 / csi1000
    p.add_argument("--markets", default="csi300,csi1000,csiall", help="comma-separated market names for D.instruments")
    p.add_argument(
        "--benchmark-map",
        default="",
        # benchmark 只影响回测基准收益/风险计算，不影响训练；但为了口径统一建议显式配置。
        help="optional overrides like `csi300:SH000300,csi1000:SH000852,all:SH000300` (default: built-in mapping)",
    )
    # label 预测目标：未来 h 日收益率（用 Qlib 表达式 Ref($close, -h)/$close-1）。
    p.add_argument("--label-horizons", default="1,5,10", help="comma-separated horizons in days")
    p.add_argument(
        "--pit-fields",
        default="assettequity_q,netprofit_q,roeavg_q,yoyni_q",
        # 注意：PIT 打开时需要你已把对应的日频 PIT 特征 dump 到 qlib bin 数据里，否则训练会直接报错。
        help="comma-separated PIT field names (used by `--pit-grid all` as the PIT-on setting)",
    )
    p.add_argument(
        "--pit-grid",
        default="all",
        choices=["none", "all", "single", "all+single"],
        # all：只做 no_pit vs pit_all（推荐默认，任务量可控）
        # single：no_pit + 每个 PIT 单字段（用于定位“哪个字段更有贡献”）
        # all+single：两者叠加（任务量最大）
        help="PIT settings to compare: none (no PIT only), all (no PIT + PIT with all --pit-fields), single (no PIT + PIT with each single field), all+single (combine both)",
    )
    p.add_argument("--start-date", default="2010-01-01")
    p.add_argument("--end-date", default="2026-01-10")

    p.add_argument("--train-years", type=int, default=3)
    p.add_argument("--valid-years", type=int, default=1)
    p.add_argument("--test-years", type=int, default=1)
    p.add_argument("--step-years", type=int, default=1)
    p.add_argument("--min-test-days", type=int, default=200)
    p.add_argument("--allow-partial-last-window", action=argparse.BooleanOptionalAction, default=True)

    p.add_argument("--exp-name-prefix", default="grid", help="experiment name will be `<prefix>_<market>`")
    # 并行度建议：
    # - workers 控制“并行 job 数”（每个 job 会跑一次训练 + 一次报告）
    # - lgb-num-threads 控制单个训练的 LightGBM 线程数
    # 两者相乘约等于 CPU 压力：建议先小再放大，否则容易互相抢资源导致整体更慢。
    # 默认 workers=2：先把流程跑稳，再按机器资源手动调大并行度。
    p.add_argument("--workers", type=int, default=2)
    p.add_argument("--lgb-num-threads", type=int, default=None, help="override trainer `--lgb-num-threads` (default: trainer default)")
    p.add_argument(
        "--joblib-backend",
        choices=["threading", "loky", "multiprocessing"],
        default="threading",
        help="passed to trainer `--joblib-backend` (default: threading)",
    )
    p.add_argument(
        "--pit-missing",
        choices=["error", "skip"],
        default="skip",
        help="for PIT-enabled jobs: how to handle instruments missing PIT bins (default: skip)",
    )
    p.add_argument(
        "--python",
        default=None,
        help="python executable for subprocesses (default: use .venv/bin/python if exists, otherwise current python)",
    )

    # Backtest/report params (kept constant for comparability).
    p.add_argument("--topk", type=int, default=10)
    p.add_argument("--n-drop", type=int, default=1)
    p.add_argument("--hold-thresh", type=int, default=5)
    p.add_argument("--account", type=float, default=100_000_000)
    p.add_argument("--open-cost", type=float, default=0.0005)
    p.add_argument("--close-cost", type=float, default=0.0015)
    p.add_argument("--min-cost", type=float, default=5.0)
    p.add_argument("--limit-threshold", type=float, default=0.095)

    p.add_argument("--out-dir", default=None, help="default: reports/grid_runs/<timestamp>")
    # resume：基于 results.jsonl 中 status=ok 的 key 跳过；可用于断点续跑（比如中途机器重启）。
    p.add_argument("--resume", action="store_true", help="skip jobs already marked ok in results.jsonl")
    p.add_argument("--dry-run", action="store_true")
    p.add_argument(
        "--export-csv",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="after run, extract Experiment Parameters + Portfolio Analysis Table into a single CSV for comparison",
    )
    p.add_argument(
        "--export-xlsx",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="also export a styled XLSX (requires openpyxl): <out_dir>/grid_compare.xlsx",
    )
    p.add_argument(
        "--progress",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="show progress bars for running jobs / exporting CSV (if tqdm is installed)",
    )
    p.add_argument(
        "--export-only",
        action="store_true",
        help="only export compare CSV from an existing --out-dir (requires --out-dir, will NOT run jobs)",
    )
    args = p.parse_args(argv)

    # 文件移动到 src/ 后，repo_root 是它的上一级目录。
    repo_root = Path(__file__).resolve().parents[1]
    if args.export_only and not args.out_dir:
        raise SystemExit("--export-only requires --out-dir")
    out_dir = Path(args.out_dir) if args.out_dir else (repo_root / "reports" / "grid_runs" / datetime.now().strftime("%Y%m%d_%H%M%S"))
    if not args.export_only:
        out_dir.mkdir(parents=True, exist_ok=True)

    # 重要：训练/报告依赖较多（plotly/statsmodels 等）。如果你在系统 python 下跑本脚本，
    # 再用 sys.executable 起子进程，很容易命中“子进程缺包”的坑。
    # 默认优先使用 repo/.venv/bin/python（若存在）。
    venv_python = repo_root / ".venv" / "bin" / "python"
    python_exe = args.python or (str(venv_python) if venv_python.exists() else sys.executable)

    results_path = out_dir / "results.jsonl"
    if args.export_only:
        if not args.export_csv and not args.export_xlsx:
            raise SystemExit("--export-only requires --export-csv and/or --export-xlsx")
        ok_rows, err_rows = _export_compare_table(repo_root=repo_root, out_dir=out_dir, results_path=results_path, progress=args.progress)
        cols = [
            "key",
            "exp_name",
            "market",
            "label_horizon",
            "label_expr",
            "pit",
            "pit_fields",
            "benchmark",
            "window.train_start",
            "window.train_end",
            "window.valid_start",
            "window.valid_end",
            "window.test_start",
            "window.test_end",
            "excess_return_without_cost.mean",
            "excess_return_without_cost.std",
            "excess_return_without_cost.annualized_return",
            "excess_return_without_cost.information_ratio",
            "excess_return_without_cost.max_drawdown",
            "excess_return_with_cost.mean",
            "excess_return_with_cost.std",
            "excess_return_with_cost.annualized_return",
            "excess_return_with_cost.information_ratio",
            "excess_return_with_cost.max_drawdown",
            "IC",
            "ICIR",
            "RankIC",
            "RankICIR",
            "pred_instruments",
            "pred_start",
            "pred_end",
            "model_class",
            "num_leaves",
            "max_depth",
            "learning_rate",
            "subsample",
            "colsample_bytree",
            "lambda_l1",
            "lambda_l2",
            "num_threads",
            "strategy.topk",
            "strategy.n_drop",
            "strategy.hold_thresh",
            "exchange.open_cost",
            "exchange.close_cost",
            "exchange.min_cost",
            "exchange.limit_threshold",
            "recorder_id",
            "report_html",
        ]
        if args.export_csv:
            out_csv = out_dir / "grid_compare.csv"
            _write_csv(out_csv, rows=ok_rows, columns=cols)
            if err_rows:
                _write_csv(out_dir / "grid_errors.csv", rows=err_rows, columns=["key", "status", "error", "train_log", "report_log"])
            print(f"[grid] exported compare csv: {out_csv}")
        if args.export_xlsx:
            try:
                out_xlsx = out_dir / "grid_compare.xlsx"
                _export_compare_xlsx(out_path=out_xlsx, rows=ok_rows, columns=cols, err_rows=err_rows, progress=args.progress)
                print(f"[grid] exported compare xlsx: {out_xlsx}")
            except Exception as exc:
                print(f"[grid] failed to export compare xlsx: {exc}")
        return 0

    start = _parse_date(args.start_date)
    end = _parse_date(args.end_date)
    markets = _comma_list(args.markets)
    horizons = [int(x) for x in _comma_list(args.label_horizons)]
    pit_fields = _comma_list(args.pit_fields)
    benchmark_map: dict[str, str] = {}
    for item in _comma_list(args.benchmark_map):
        if ":" not in item:
            continue
        k, v = item.split(":", 1)
        k = k.strip().lower()
        v = v.strip()
        if k and v:
            benchmark_map[k] = v

    windows = _iter_walk_forward_windows(
        start=start,
        end=end,
        train_years=args.train_years,
        valid_years=args.valid_years,
        test_years=args.test_years,
        step_years=args.step_years,
        allow_partial_last_window=args.allow_partial_last_window,
        min_test_days=args.min_test_days,
    )
    if not windows:
        raise SystemExit("no walk-forward windows generated; adjust start/end or year settings")

    pit_specs: list[PitSpec] = [PitSpec(name="no_pit", pit_fields_arg=["--pit-fields"])]
    pit_fields_raw = args.pit_fields.strip()
    if args.pit_grid in ("all", "all+single"):
        if pit_fields_raw:
            pit_specs.append(PitSpec(name="pit_all", pit_fields_arg=["--pit-fields", pit_fields_raw]))
    if args.pit_grid in ("single", "all+single"):
        for f in pit_fields:
            pit_specs.append(PitSpec(name=f"pit_{f}", pit_fields_arg=["--pit-fields", f]))

    jobs: list[JobSpec] = []
    for market in markets:
        exp_name = f"{_slug(args.exp_name_prefix)}_{_slug(market)}"
        bench = benchmark_map.get(market.lower(), _benchmark_for_market(market))
        for h in horizons:
            label_expr = f"Ref($close, -{h}) / $close - 1"
            for pit in pit_specs:
                for w in windows:
                    jobs.append(
                        JobSpec(
                            market=market,
                            benchmark=bench,
                            label_horizon=h,
                            label_expr=label_expr,
                            pit=pit,
                            window=w,
                            exp_name=exp_name,
                        )
                    )

    completed = _load_completed_keys(results_path) if args.resume else set()
    todo = [j for j in jobs if j.tag not in completed]

    summary = {
        "provider_uri": args.provider_uri,
        "start_date": args.start_date,
        "end_date": args.end_date,
        "markets": markets,
        "benchmark_map": benchmark_map,
        "label_horizons": horizons,
        "pit_specs": [ps.name for ps in pit_specs],
        "windows": [asdict(w) for w in windows],
        "num_jobs": len(jobs),
        "num_todo": len(todo),
        "out_dir": str(out_dir),
        "workers": args.workers,
        "lgb_num_threads": args.lgb_num_threads,
        "python": python_exe,
        "pit_missing": args.pit_missing,
        "joblib_backend": args.joblib_backend,
    }
    (out_dir / "summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))

    if args.dry_run:
        # dry-run 只负责展开网格与窗口，便于你先评估任务量与时间区间是否符合预期。
        return 0

    lock = threading.Lock()

    def write_result(res: JobResult) -> None:
        # 采用 jsonl：每完成一个 job 立即 append 一行，抗中断（不会等全部结束才写）。
        row = {
            "key": res.job.tag,
            "status": res.status,
            "started_at": res.started_at,
            "finished_at": res.finished_at,
            "market": res.job.market,
            "benchmark": res.job.benchmark,
            "label_horizon": res.job.label_horizon,
            "label_expr": res.job.label_expr,
            "pit": res.job.pit.name,
            "window": asdict(res.job.window),
            "exp_name": res.job.exp_name,
            "recorder_id": res.recorder_id,
            "report_html": res.report_html,
            "train_returncode": res.train_returncode,
            "report_returncode": res.report_returncode,
            "error": res.error,
            "train_log": res.train_log,
            "report_log": res.report_log,
        }
        line = json.dumps(row, ensure_ascii=False)
        with lock:
            results_path.parent.mkdir(parents=True, exist_ok=True)
            with results_path.open("a", encoding="utf-8") as f:
                f.write(line + "\n")

    # Use threads: each job runs external python processes; easy concurrency without sharing Qlib state.
    from concurrent.futures import ThreadPoolExecutor, as_completed

    print(f"[grid] running jobs: {len(todo)} (workers={args.workers})")
    with ThreadPoolExecutor(max_workers=args.workers) as ex:
        futs = []
        for job in todo:
            futs.append(
                ex.submit(
                    _run_one_job,
                    job=job,
                    repo_root=repo_root,
                    python_exe=python_exe,
                    provider_uri=args.provider_uri,
                    lgb_num_threads=args.lgb_num_threads,
                    topk=args.topk,
                    n_drop=args.n_drop,
                    hold_thresh=args.hold_thresh,
                    account=args.account,
                    open_cost=args.open_cost,
                    close_cost=args.close_cost,
                    min_cost=args.min_cost,
                    limit_threshold=args.limit_threshold,
                    out_dir=out_dir,
                    pit_missing=args.pit_missing,
                    joblib_backend=args.joblib_backend,
                )
            )

        ok = 0
        failed = 0
        bar = None
        if args.progress and tqdm is not None:
            bar = tqdm(total=len(futs), desc="grid jobs", dynamic_ncols=True)  # type: ignore[assignment]

        def _log(msg: str) -> None:
            if bar is not None:
                try:
                    bar.write(msg)  # type: ignore[union-attr]
                    return
                except Exception:
                    pass
            print(msg)

        for fut in as_completed(futs):
            res = fut.result()
            write_result(res)
            if res.status == "ok":
                ok += 1
                _log(f"[ok] {res.job.tag} recorder={res.recorder_id} html={res.report_html}")
            else:
                failed += 1
                # report_failed 时优先提示 report.log；train_failed 则提示 train.log
                log_path = res.report_log if res.status.startswith("report") else res.train_log
                _log(f"[fail] {res.job.tag} status={res.status} log={log_path}")
            if bar is not None:
                try:
                    bar.update(1)  # type: ignore[union-attr]
                    bar.set_postfix(ok=ok, failed=failed)  # type: ignore[union-attr]
                except Exception:
                    pass
        if bar is not None:
            try:
                bar.close()  # type: ignore[union-attr]
            except Exception:
                pass
        print(f"[grid] done ok={ok} failed={failed} results={results_path}")

    if args.export_csv or args.export_xlsx:
        try:
            ok_rows, err_rows = _export_compare_table(repo_root=repo_root, out_dir=out_dir, results_path=results_path, progress=args.progress)
            cols = [
                "key",
                "exp_name",
                "market",
                "label_horizon",
                "label_expr",
                "pit",
                "pit_fields",
                "benchmark",
                "window.train_start",
                "window.train_end",
                "window.valid_start",
                "window.valid_end",
                "window.test_start",
                "window.test_end",
                "excess_return_without_cost.mean",
                "excess_return_without_cost.std",
                "excess_return_without_cost.annualized_return",
                "excess_return_without_cost.information_ratio",
                "excess_return_without_cost.max_drawdown",
                "excess_return_with_cost.mean",
                "excess_return_with_cost.std",
                "excess_return_with_cost.annualized_return",
                "excess_return_with_cost.information_ratio",
                "excess_return_with_cost.max_drawdown",
                "IC",
                "ICIR",
                "RankIC",
                "RankICIR",
                "pred_instruments",
                "pred_start",
                "pred_end",
                "model_class",
                "num_leaves",
                "max_depth",
                "learning_rate",
                "subsample",
                "colsample_bytree",
                "lambda_l1",
                "lambda_l2",
                "num_threads",
                "strategy.topk",
                "strategy.n_drop",
                "strategy.hold_thresh",
                "exchange.open_cost",
                "exchange.close_cost",
                "exchange.min_cost",
                "exchange.limit_threshold",
                "recorder_id",
                "report_html",
            ]
            if args.export_csv:
                out_csv = out_dir / "grid_compare.csv"
                _write_csv(out_csv, rows=ok_rows, columns=cols)
                if err_rows:
                    _write_csv(out_dir / "grid_errors.csv", rows=err_rows, columns=["key", "status", "error", "train_log", "report_log"])
                print(f"[grid] exported compare csv: {out_csv}")
            if args.export_xlsx:
                try:
                    out_xlsx = out_dir / "grid_compare.xlsx"
                    _export_compare_xlsx(out_path=out_xlsx, rows=ok_rows, columns=cols, err_rows=err_rows, progress=args.progress)
                    print(f"[grid] exported compare xlsx: {out_xlsx}")
                except Exception as exc:
                    print(f"[grid] failed to export compare xlsx: {exc}")
        except Exception as exc:
            print(f"[grid] failed to export compare table: {exc}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
